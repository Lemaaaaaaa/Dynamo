import clr
import openpyxl

# Inputs from Dynamo
excel_file = IN[0]  # User-defined file path
sheet_name = IN[1]  # User-defined sheet name

try:
    # Load the Excel workbook
    workbook = openpyxl.load_workbook(excel_file, data_only=True)  # data_only=True reads cell values (not formulas)
    
    # Check if sheet exists
    if sheet_name not in workbook.sheetnames:
        OUT = f"❌ Error: Sheet '{sheet_name}' not found in the file."
    else:
        sheet = workbook[sheet_name]  # Get the specified sheet

        # Extract all rows and columns into a list of lists
        excel_data = []
        for row in sheet.iter_rows(values_only=True):
            excel_data.append([cell if cell is not None else None for cell in row])  # Replace empty cells with None

        # Output the extracted data
        OUT = excel_data

except Exception as e:
    OUT = f"❌ Error: {e}"
    
    
    # Input: Bar data from Dynamo (including Bar ID and Group Name)
bar_data = IN[0]  # List of lists, where last index contains the group name

# Step 1: Create a dictionary to store grouped bars
group_dict = {}

# Step 2: Loop through each bar entry
for bar in bar_data:
    try:
        bar_id = bar[0]  # Extract Bar ID
        group_name = bar[-1]  # Extract Group Name (last column)

        # Convert to string first, in case it's a number
        group_name = str(group_name).strip()

        # Ignore empty group names
        if group_name == "":
            continue

        # Normalize group name to lowercase (only if it's a string)
        group_name = group_name.lower()

        # Add bar to the corresponding group
        if group_name not in group_dict:
            group_dict[group_name] = []
        group_dict[group_name].append(bar_id)

    except Exception as e:
        print(f"❌ Error processing bar {bar}: {e}")

# Step 3: Convert dictionary to a list of lists
grouped_bar_list = [[group, bars] for group, bars in group_dict.items()]

# Output to Dynamo
OUT = grouped_bar_list
