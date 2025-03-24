# Load the Python Standard and DesignScript Libraries
import sys
import clr
clr.AddReference('ProtoGeometry')
from Autodesk.DesignScript.Geometry import *
import sys
import clr
import csv
import os
import hashlib
import openpyxl


bar_data = IN[0]
sheet_name = IN[1]  # Change this to the correct sheet name
if IN[2]: # Button(boolean)
    # Define file path
    excel_file = r"C:\Documents\DYNAMO TRIAL\Grouping\Test.xlsx"
    
    # Column headers
    headers = ["Panel", "FaceAsssignment", "Bar ID", "startNodeId", "startNodeIdX", "startNodeIdY", "startNodeIdZ", "endNodeId", "endNodeIdX", "endNodeIdY", "endNodeIdZ", "Length", "Section_Name", "Material", "sectionWidth", "ShapeID", "sectionPerimeter", "nominalWeight", "Af", "Ac"]
    
    # Function to compute hash of current bar data
    def compute_data_hash(data):
        data_string = "".join([",".join(map(str, row)) for row in data])  # Convert list to string
        return hashlib.md5(data_string.encode()).hexdigest()  # Compute hash
    
    # Read existing file (if it exists) and compute its hash
    previous_hash = None
    if os.path.exists(excel_file):
        # Load the existing Excel file
        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook[sheet_name]
    
        # Read existing data as a string
        existing_data = []
        for row in sheet.iter_rows(values_only=True):
            existing_data.append(row)
        
        existing_data_str = "".join(map(str, existing_data))
        previous_hash = hashlib.md5(existing_data_str.encode()).hexdigest()
    
    # Compute hash of new data
    new_hash = compute_data_hash(bar_data)
    
    # Check if data has changed
    if new_hash != previous_hash:
        print("Changes detected: Updating Excel file...")
    
        # If file exists, clear the sheet before writing
        if os.path.exists(excel_file):
            workbook = openpyxl.load_workbook(excel_file)
            if sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                sheet.delete_rows(1, sheet.max_row)  # Deletes all rows
            else:
                sheet = workbook.create_sheet(sheet_name)
        else:
            # Create a new workbook and sheet
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = sheet_name
    
        # Write headers
        sheet.append(headers)
    
        # Write new data
        for row in bar_data:
            sheet.append(row)
    
        # Save the updated file
        workbook.save(excel_file)
        OUT= f"Excel file updated: {excel_file}"
    
    else:
        OUT= "No changes detected: Excel file remains unchanged.")
    
    # Output file path for Dynamo
    OUT = excel_file
