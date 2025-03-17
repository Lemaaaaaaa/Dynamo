import sys
import clr
import csv
import os
import hashlib
import openpyxl


bars = IN[0]
sheet_name = IN[1]  # Change this to the correct sheet name


# Place your code below this line
bar_data = []
for bar in bars:
    try:
        # Extract bar information
        bar_id = bar.Id
        start_id = bar.StartNode.Id
        start_x = bar.StartNode.X
        start_y = bar.StartNode.Y
        start_z = bar.StartNode.Z
        end_id = bar.EndNode.Id
        end_x = bar.EndNode.X
        end_y = bar.EndNode.Y
        end_z = bar.EndNode.Z

        # Append as a list (row) to bar_data
        bar_data.append([bar_id, start_id, start_x, start_y, start_z, end_id, end_x, end_y, end_z])

    except Exception as e:
        bar_data.append([f"Error processing bar {bar.Id}", str(e)])
# Define file path
excel_file = r"C:\Users\BCS\BSSE\BSS BIM - Documents\DYNAMO TRIAL\TowerToRobot_B.xlsx"

# Column headers
headers = ["Bar ID", "Start Node ID", "Start X", "Start Y", "Start Z", "End Node ID", "End X", "End Y", "End Z"]

# Function to compute hash of current bar data
def compute_data_hash(data):
    data_string = "".join([",".join(map(str, row)) for row in data])  # Convert list to string
    return hashlib.md5(data_string.encode()).hexdigest()  # Compute hash

# Read existing file (if it exists) and compute its hash
previous_hash = None
if os.path.exists(excel_file):
    # Load the existing Excel file
    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook[sheet_name]

    # Read existing data as a string
    existing_data = []
    for row in sheet.iter_rows(values_only=True):
        existing_data.append(row)
    
    existing_data_str = "".join(map(str, existing_data))
    previous_hash = hashlib.md5(existing_data_str.encode()).hexdigest()

# Compute hash of new data
new_hash = compute_data_hash(bar_data)

# Check if data has changed
if new_hash != previous_hash:
    print("🔄 Changes detected: Updating Excel file...")

    # If file exists, clear the sheet before writing
    if os.path.exists(excel_file):
        workbook = openpyxl.load_workbook(excel_file)
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            sheet.delete_rows(1, sheet.max_row)  # Deletes all rows
        else:
            sheet = workbook.create_sheet(sheet_name)
    else:
        # Create a new workbook and sheet
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = sheet_name

    # Write headers
    sheet.append(headers)

    # Write new data
    for row in bar_data:
        sheet.append(row)

    # Save the updated file
    workbook.save(excel_file)
    print(f"✅ Excel file updated: {excel_file}")

else:
    print("✅ No changes detected: Excel file remains unchanged.")

# Output file path for Dynamo
OUT = excel_file